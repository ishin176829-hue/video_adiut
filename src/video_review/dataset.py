from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from .config import settings
from .policies import keyword_findings, load_policy
from .utils import new_id


EPISODE_REASON_RE = re.compile(r"【集名称：([^】]+)】([^【]*)")


def _records_from_rows(rows: list[dict]) -> list[dict]:
    records = []
    policy = load_policy()
    for row in rows:
        drama_id = str(row.get("剧ID") or row.get("drama_id") or "")
        drama_name = str(row.get("剧名") or row.get("drama_name") or "")
        video_url = str(row.get("视频链接") or row.get("video_url") or row.get("url") or "")
        text = "\n".join(str(row.get(k) or "") for k in ["拒绝理由汇总_1", "拒绝理由汇总_2", "reject_reason", "reason"])
        matches = EPISODE_REASON_RE.findall(text)
        if not matches and text.strip():
            matches = [(str(row.get("集名称") or row.get("episode_name") or ""), text)]
        for episode_name, reason in matches:
            reason = " ".join(reason.split())
            if not reason:
                continue
            findings = keyword_findings(reason, policy)
            records.append(
                {
                    "drama_id": drama_id,
                    "drama_name": drama_name,
                    "episode_name": episode_name,
                    "video_url": video_url,
                    "reject_reason": reason,
                    "matched_categories": [f.category for f in findings],
                }
            )
    return records


def import_xlsx(path: Path) -> Path:
    settings.ensure_dirs()
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        values = ws.iter_rows(values_only=True)
        header = next(values, None)
        if not header:
            continue
        keys = [str(v).strip() if v is not None else "" for v in header]
        for row in values:
            data = {keys[i]: row[i] for i in range(min(len(keys), len(row))) if keys[i]}
            if any(v is not None and str(v).strip() for v in data.values()):
                rows.append(data)
    return write_records(_records_from_rows(rows))


def import_csv(path: Path) -> Path:
    settings.ensure_dirs()
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    return write_records(_records_from_rows(rows))


def write_records(records: list[dict]) -> Path:
    output = settings.datasets_dir / f"rejected_dataset_{new_id('batch')}.jsonl"
    with output.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return output
